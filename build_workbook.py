import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial", color="000000")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(name="Arial", bold=True)
CAMPAIGN_FILL = PatternFill("solid", fgColor="BDD7EE")
CAMPAIGN_FONT = Font(name="Arial", bold=True)
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY = '$#,##0.00;($#,##0.00);"-"'

wb = openpyxl.Workbook()
wb.remove(wb.active)

def style_header(ws, row, ncols, height=30):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def border_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).border = BORDER

# =================================================================
# READ ME
# =================================================================
rm = wb.create_sheet("Read Me")
set_widths(rm, [38, 100])
rm["A1"] = "Revenue Recognition Workbook — How This Is Organized"
rm["A1"].font = TITLE_FONT
rm.merge_cells("A1:B1")

r = 3
def add_line(label, text, bold_label=True, fill=None):
    global r
    rm.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=bold_label)
    cell = rm.cell(row=r, column=2, value=text)
    cell.font = BLACK
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if fill:
        rm.cell(row=r, column=1).fill = fill
        cell.fill = fill
    rm.row_dimensions[r].height = 30
    r += 1

add_line("Tab", "Purpose")
rm.cell(row=r-1, column=1).font = SUBHEADER_FONT
rm.cell(row=r-1, column=2).font = SUBHEADER_FONT
add_line("Campaigns", "One row per overarching campaign budget. Shows the total budget, how much is allocated to sub-projects/SOWs so far, how much is earned/open, and — for flex budgets — how much is still available to move to a different sub-project within that same campaign.")
add_line("Sub-Projects (SOWs)", "One row per SOW or task line, each tied to a Campaign ID. Earned-to-date and Open pull automatically from the Monthly Recognition tab, so you no longer total months by hand.")
add_line("Monthly Recognition", "One row per month per SOW (long format instead of a 20-column-wide grid). Adding a new month is just adding new rows — no need to insert columns across every campaign block. Each row also tags the invoice number tied to that revenue.")
add_line("Invoices - Finance", "The list you hand to Finance: invoice number, campaign, SOW, amount, and status (sent / not sent / confirmed). Filter to 'Not Yet Sent' each period to see exactly what still needs to go to Finance.")

r += 1
add_line("Color key", "")
rm.cell(row=r-1, column=1).font = SUBHEADER_FONT
rm.cell(row=r-1, column=2).font = SUBHEADER_FONT
rm.cell(row=r, column=1, value="Blue text").font = BLUE
rm.cell(row=r, column=2, value="Editable input — type or change these directly.").font = BLACK
r += 1
rm.cell(row=r, column=1, value="Black text").font = BLACK
rm.cell(row=r, column=2, value="Formula — calculated automatically, don't overwrite.").font = BLACK
r += 1
rm.cell(row=r, column=1, value="Yellow highlight").fill = FLAG_FILL
rm.cell(row=r, column=2, value="Data-quality flag carried over from your original file — needs a decision from you.").font = BLACK
r += 2

add_line("Data-quality flags found in your original file", "")
rm.cell(row=r-1, column=1).font = SUBHEADER_FONT
rm.cell(row=r-1, column=2).font = SUBHEADER_FONT
flags = [
    ("Duplicate Campaign ID", "\"Campaign UCMC-010\" was used for both Strategic Counsel & Advisory Support and Comms AOR. Split into CMP-06 and CMP-07 here — confirm whether they should actually share one flex pool or stay separate."),
    ("Comms AOR has no top-level budget", "The original file never set an Original/Revised Budget for the Comms AOR campaign header. This workbook backed into $240,000 from the known monthly invoices — confirm the real approved budget."),
    ("Landing Page Design (26-UCMC-050) has no budget", "Only a $1,032.50 earned amount exists for this SOW; Original Budget was blank in your file. Flagged for you to fill in."),
    ("OOP Expense line with no invoice number", "Comms AOR's May OOP Expense ($6,278.58) has no invoice reference in the source file, unlike the March OOP line which cites INV-39804AO."),
    ("Service Line Campaigns has no sub-projects", "The campaign header exists with a $0 budget but no SOWs were ever itemized under it."),
    ("Broken reference errors in original file", "Cancer Center Pavilion and Haas F1 Partnership had broken-reference (REF) errors in the Revised Budget column of your original workbook — rebuilt from scratch here."),
    ("Most revenue has no invoice # tied to it", "Only Haas F1 Partnership and Comms AOR embed invoice numbers in your original file. The other 5 campaigns have never had an invoice number captured against the revenue — see the Invoices - Finance tab."),
]
for label, text in flags:
    add_line(label, text, fill=FLAG_FILL)
rm.sheet_view.showGridLines = False

# =================================================================
# CAMPAIGNS
# =================================================================
camp = wb.create_sheet("Campaigns")
set_widths(camp, [11, 38, 10, 16, 16, 20, 20, 16, 16, 46])
headers = ["Campaign ID", "Campaign Name", "Flex?", "Original Budget", "Revised Budget",
           "Allocated to Sub-Projects", "Available to Reallocate (Flex)", "Total Earned",
           "Total Open", "Notes"]
for i, h in enumerate(headers, start=1):
    camp.cell(row=1, column=i, value=h)
style_header(camp, 1, len(headers))
camp.freeze_panes = "A2"

campaigns = [
    ("CMP-01", "Brand Campaign Creative", "Y", 2015000, 2015000, "Original UCMC-004."),
    ("CMP-02", "Service Line Campaigns", "Y", 0, 0, "UCMC-006. No sub-projects defined yet in source file."),
    ("CMP-03", "Cancer Center Pavilion", "Y", 2215000, 2215000, "Includes standalone SOW 25-UCMC-200 plus the UCMC-007 sub-project group. Original file had no usable top-level budget (broken reference error) — this total is reconstructed as the sum of its sub-projects. Confirm the real approved campaign budget."),
    ("CMP-04", "U.S. News & World Report Voting/Reputation", "Y", 528000, 528000, ""),
    ("CMP-05", "Haas F1 Partnership", "Y", 247250, 247250, "UCMC-009. Invoiced monthly."),
    ("CMP-06", "Strategic Counsel & Advisory Support", "Y", 490000, 490000, "UCMC-010."),
    ("CMP-07", "Comms AOR", "Y", 255153.5, 255153.5, "Also referenced as UCMC-010 in source file (duplicate ID) — no top-level budget was set originally; reconstructed as the sum of known invoices. Confirm real contracted budget."),
]

row = 2
for cid, name, flex, orig, rev, note in campaigns:
    camp.cell(row=row, column=1, value=cid).font = BLACK
    camp.cell(row=row, column=2, value=name).font = BLACK
    camp.cell(row=row, column=3, value=flex).font = BLUE
    camp.cell(row=row, column=4, value=orig).font = BLUE
    camp.cell(row=row, column=4).number_format = CURRENCY
    camp.cell(row=row, column=5, value=rev).font = BLUE
    camp.cell(row=row, column=5).number_format = CURRENCY
    camp.cell(row=row, column=6, value=f"=SUMIF('Sub-Projects (SOWs)'!$B$2:$B$40,$A{row},'Sub-Projects (SOWs)'!$F$2:$F$40)").font = BLACK
    camp.cell(row=row, column=6).number_format = CURRENCY
    camp.cell(row=row, column=7, value=f"=E{row}-F{row}").font = BLACK
    camp.cell(row=row, column=7).number_format = CURRENCY
    camp.cell(row=row, column=8, value=f"=SUMIF('Sub-Projects (SOWs)'!$B$2:$B$40,$A{row},'Sub-Projects (SOWs)'!$G$2:$G$40)").font = BLACK
    camp.cell(row=row, column=8).number_format = CURRENCY
    camp.cell(row=row, column=9, value=f"=E{row}-H{row}").font = BLACK
    camp.cell(row=row, column=9).number_format = CURRENCY
    camp.cell(row=row, column=10, value=note).font = Font(name="Arial", size=9, italic=True, color="808080")
    if "Confirm" in note or "flag" in note.lower():
        for c in range(1, 11):
            camp.cell(row=row, column=c).fill = FLAG_FILL
    border_row(camp, row, 10)
    row += 1

total_row = row
camp.cell(row=total_row, column=2, value="TOTAL").font = Font(name="Arial", bold=True)
for col in [4, 5, 6, 7, 8, 9]:
    letter = get_column_letter(col)
    camp.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{row-1})").font = Font(name="Arial", bold=True)
    camp.cell(row=total_row, column=col).number_format = CURRENCY
border_row(camp, total_row, 10)
camp.sheet_view.showGridLines = False

# =================================================================
# SUB-PROJECTS (SOWs)
# =================================================================
sp = wb.create_sheet("Sub-Projects (SOWs)")
set_widths(sp, [16, 11, 14, 42, 16, 16, 16, 16, 14, 40])
sp_headers = ["SOW/Task ID", "Campaign ID", "Legacy Group Ref", "Description",
              "Original Budget", "Revised Budget", "Earned to Date", "Open", "Status", "Notes"]
for i, h in enumerate(sp_headers, start=1):
    sp.cell(row=1, column=i, value=h)
style_header(sp, 1, len(sp_headers))
sp.freeze_panes = "A2"

# (sow_id, campaign_id, legacy_group, description, orig_budget, status, note, flag)
sow_rows = [
    ("25-UCMC-016", "CMP-01", "UCMC-004", "Branding Campaign - Concepting", 300000, "In Progress", "", False),
    ("25-UCMC-018", "CMP-01", "UCMC-004", "Forecasted - Brand Campaign Deliverables", 700000, "Not Started", "", False),
    ("25-UCMC-019", "CMP-01", "UCMC-004", "Branding Campaign - Video/Photoshoot", 1000000, "Not Started", "", False),
    ("25-UCMC-023", "CMP-01", "UCMC-004", "Brand Campaign Concept Testing", 15000, "Complete", "", False),

    ("25-UCMC-200", "CMP-03", "(direct)", "Cancer Pavilion Plan", 125000, "In Progress", "", False),
    ("26-UCMC-047", "CMP-03", "UCMC-007", "Creative Concepting", 175000, "In Progress", "", False),
    ("CMP3-EXTRA1", "CMP-03", "UCMC-007", "Creative Development & Production", 450000, "Not Started", "No SOW code in source file.", False),
    ("26-UCMC-050", "CMP-03", "UCMC-007", "Landing Page Design", 0, "In Progress", "No Original Budget in source file — confirm amount.", True),
    ("26-UCMC-036", "CMP-03", "UCMC-007", "Integrated Engagement Strategy & Plan", 85000, "In Progress", "", False),
    ("26-UCMC-037", "CMP-03", "UCMC-007", "Program Governance Meetings", 65000, "In Progress", "", False),
    ("26-UCMC-040", "CMP-03", "UCMC-007", "SEO Strategy & Optimization", 50000, "Not Started", "", False),
    ("26-UCMC-041", "CMP-03", "UCMC-007", "Earned Media", 25000, "Not Started", "", False),
    ("26-UCMC-038", "CMP-03", "UCMC-007", "Paid Media Expenses", 434625, "Not Started", "", False),
    ("26-UCMC-039", "CMP-03", "UCMC-007", "Paid Media Labor & Measurement", 40375, "Not Started", "", False),
    ("26-UCMC-042", "CMP-03", "UCMC-007", "Event Planning", 250000, "In Progress", "", False),
    ("26-UCMC-043", "CMP-03", "UCMC-007", "Event Travel OOP Expenses", 15000, "Not Started", "", False),
    ("26-UCMC-044", "CMP-03", "UCMC-007", "Community Engagement", 50000, "In Progress", "", False),
    ("26-UCMC-029", "CMP-03", "UCMC-007", "Account Management & Advisory", 225000, "In Progress", "", False),
    ("CMP3-EXTRA2", "CMP-03", "UCMC-007", "Event Production & Vendor Estimated Hard Cost", 225000, "Not Started", "No SOW code in source file.", False),

    ("25-UCMC-021", "CMP-04", "(direct)", "U.S. News & World Report Voting Campaign", 440000, "Complete", "", False),
    ("26-UCMC-035", "CMP-04", "(direct)", "UChicago National Reputation Campaign - Conference Support", 88000, "In Progress", "", False),

    ("CMP5-SOW1", "CMP-05", "UCMC-009", "Haas F1 Partnership: Overall Strategic Support", 247250, "In Progress", "Invoiced monthly — see Invoices - Finance tab for the 12 monthly invoice lines.", False),

    ("25-UCMC-001", "CMP-06", "UCMC-010", "UChicago Account Oversight", 310000, "In Progress", "", False),
    ("CMP6-EXTRA1", "CMP-06", "UCMC-010", "OOP Expenses", 30000, "In Progress", "No SOW code in source file.", False),
    ("CMP6-EXTRA2", "CMP-06", "UCMC-010", "Annual Planning", 150000, "Not Started", "", False),

    ("CMP7-SOW1", "CMP-07", "UCMC-010 (dup)", "Comms AOR Retainer", 240000, "In Progress", "Budget backed into from known monthly invoices — confirm actual contracted amount.", True),
    ("CMP7-EXTRA1", "CMP-07", "UCMC-010 (dup)", "OOP Expenses", 15153.5, "In Progress", "Budget backed into from known monthly invoices — confirm actual contracted amount.", True),
]

row = 2
sow_row_map = {}
for sow_id, cid, legacy, desc, orig, status, note, flag in sow_rows:
    sow_row_map[sow_id] = row
    sp.cell(row=row, column=1, value=sow_id).font = BLACK
    sp.cell(row=row, column=2, value=cid).font = BLUE
    sp.cell(row=row, column=3, value=legacy).font = Font(name="Arial", size=9, color="808080")
    sp.cell(row=row, column=4, value=desc).font = BLACK
    sp.cell(row=row, column=5, value=orig).font = BLUE
    sp.cell(row=row, column=5).number_format = CURRENCY
    sp.cell(row=row, column=6, value=orig).font = BLUE
    sp.cell(row=row, column=6).number_format = CURRENCY
    sp.cell(row=row, column=7, value=f"=SUMIF('Monthly Recognition'!$A$2:$A$100,$A{row},'Monthly Recognition'!$E$2:$E$100)").font = BLACK
    sp.cell(row=row, column=7).number_format = CURRENCY
    sp.cell(row=row, column=8, value=f"=F{row}-G{row}").font = BLACK
    sp.cell(row=row, column=8).number_format = CURRENCY
    sp.cell(row=row, column=9, value=status).font = BLUE
    sp.cell(row=row, column=10, value=note).font = Font(name="Arial", size=9, italic=True, color="808080")
    if flag:
        for c in range(1, 11):
            sp.cell(row=row, column=c).fill = FLAG_FILL
    border_row(sp, row, 10)
    row += 1
sp_last_row = row - 1
sp.sheet_view.showGridLines = False

# =================================================================
# MONTHLY RECOGNITION (tidy/long format)
# =================================================================
mr = wb.create_sheet("Monthly Recognition")
set_widths(mr, [16, 13, 10, 10, 16, 18, 12, 34])
mr_headers = ["SOW/Task ID", "Campaign ID", "Month", "Type", "Amount", "Invoice #", "Recognized?", "Notes"]
for i, h in enumerate(mr_headers, start=1):
    mr.cell(row=1, column=i, value=h)
style_header(mr, 1, len(mr_headers))
mr.freeze_panes = "A2"

# (sow_id, month, type, amount, invoice, note)
mr_rows = [
    ("25-UCMC-016", "Nov-25", "Fee", 75003.0, "Not yet assigned", ""),
    ("25-UCMC-016", "Dec-25", "Fee", 75000.0, "Not yet assigned", ""),
    ("25-UCMC-016", "Jan-26", "Fee", 70746.75, "Not yet assigned", ""),
    ("25-UCMC-016", "Feb-26", "Fee", 600.0, "Not yet assigned", ""),
    ("25-UCMC-023", "Dec-25", "Fee", 9000.0, "Not yet assigned", ""),
    ("25-UCMC-023", "Jan-26", "Fee", 6000.0, "Not yet assigned", ""),

    ("25-UCMC-200", "Nov-25", "Fee", 62500.0, "Not yet assigned", ""),
    ("25-UCMC-200", "Dec-25", "Fee", 62500.0, "Not yet assigned", ""),
    ("26-UCMC-047", "Apr-26", "Fee", 1443.75, "Not yet assigned", ""),
    ("26-UCMC-047", "May-26", "Fee", 52907.5, "Not yet assigned", ""),
    ("26-UCMC-050", "May-26", "Fee", 1032.5, "Not yet assigned", ""),
    ("26-UCMC-036", "Apr-26", "Fee", 246.25, "Not yet assigned", ""),
    ("26-UCMC-036", "May-26", "Fee", 9755.0, "Not yet assigned", ""),
    ("26-UCMC-037", "May-26", "Fee", 970.0, "Not yet assigned", ""),
    ("26-UCMC-042", "May-26", "Fee", 4342.5, "Not yet assigned", ""),
    ("26-UCMC-044", "Apr-26", "Fee", 350.0, "Not yet assigned", ""),
    ("26-UCMC-044", "May-26", "Fee", 3146.25, "Not yet assigned", ""),
    ("26-UCMC-029", "Apr-26", "Fee", 33515.0, "Not yet assigned", ""),
    ("26-UCMC-029", "May-26", "Fee", 23377.5, "Not yet assigned", ""),

    ("25-UCMC-021", "Dec-25", "Fee", 70000.0, "Not yet assigned", ""),
    ("25-UCMC-021", "Jan-26", "Fee", 200000.0, "Not yet assigned", ""),
    ("25-UCMC-021", "Feb-26", "Fee", 110000.0, "Not yet assigned", ""),
    ("25-UCMC-021", "Mar-26", "Fee", 60000.0, "Not yet assigned", ""),
    ("26-UCMC-035", "Apr-26", "Fee", 22000.0, "Not yet assigned", ""),
    ("26-UCMC-035", "May-26", "Fee", 52000.0, "Not yet assigned", ""),

    ("CMP5-SOW1", "Jan-26", "Fee", 19500.0, "INV-38952AF", "Jan Advisory/shot list"),
    ("CMP5-SOW1", "Feb-26", "Fee", 14500.0, "INV-38953AF", "Feb Advisory/Landing pg/Social"),
    ("CMP5-SOW1", "Apr-26", "Fee", 30000.0, "INV-38953AF", "Remainder of Feb invoice, recognized in April per source file"),
    ("CMP5-SOW1", "Mar-26", "Fee", 14500.0, "INV-39386AF", "March"),
    ("CMP5-SOW1", "Apr-26", "Fee", 21500.0, "INV-39927AF", "April Advisory"),
    ("CMP5-SOW1", "May-26", "Fee", 22500.0, "INV-40021AF", "May Advisory"),

    ("25-UCMC-001", "Nov-25", "Fee", 15000.0, "Not yet assigned", ""),
    ("25-UCMC-001", "Dec-25", "Fee", 15000.0, "Not yet assigned", ""),
    ("25-UCMC-001", "Jan-26", "Fee", 20000.0, "Not yet assigned", ""),
    ("25-UCMC-001", "Feb-26", "Fee", 20000.0, "Not yet assigned", ""),
    ("25-UCMC-001", "Mar-26", "Fee", 15000.0, "Not yet assigned", ""),
    ("25-UCMC-001", "Apr-26", "Fee", 15000.0, "Not yet assigned", ""),
    ("25-UCMC-001", "May-26", "Fee", 15000.0, "Not yet assigned", ""),
    ("CMP6-EXTRA1", "Apr-26", "Expense", 650.0, "Not yet assigned", "OOP Expenses"),
    ("CMP6-EXTRA1", "May-26", "Expense", 509.3, "Not yet assigned", "OOP Expenses"),

    ("CMP7-SOW1", "Mar-26", "Fee", 60000.0, "INV-37370AF", ""),
    ("CMP7-SOW1", "Apr-26", "Fee", 60000.0, "INV-37371AF", ""),
    ("CMP7-SOW1", "May-26", "Fee", 60000.0, "INV-37372AF", ""),
    ("CMP7-EXTRA1", "Mar-26", "Expense", 8874.92, "INV-39804AO", "OOP Expenses"),
    ("CMP7-EXTRA1", "May-26", "Fee", 6278.58, "Not yet assigned", "OOP Expenses — no invoice # in source file"),
]

row = 2
for sow_id, month, typ, amount, invoice, note in mr_rows:
    cid_row = sow_row_map.get(sow_id)
    mr.cell(row=row, column=1, value=sow_id).font = BLUE
    if cid_row:
        mr.cell(row=row, column=2, value=f"=INDEX('Sub-Projects (SOWs)'!$B$2:$B$40,MATCH($A{row},'Sub-Projects (SOWs)'!$A$2:$A$40,0))").font = BLACK
    mr.cell(row=row, column=3, value=month).font = BLUE
    mr.cell(row=row, column=4, value=typ).font = BLUE
    mr.cell(row=row, column=5, value=amount).font = BLUE
    mr.cell(row=row, column=5).number_format = CURRENCY
    mr.cell(row=row, column=6, value=invoice).font = (Font(name="Arial", color="C00000") if invoice == "Not yet assigned" else BLUE)
    mr.cell(row=row, column=7, value="Y").font = BLUE
    mr.cell(row=row, column=8, value=note).font = Font(name="Arial", size=9, italic=True, color="808080")
    if invoice == "Not yet assigned":
        mr.cell(row=row, column=6).fill = FLAG_FILL
    border_row(mr, row, 8)
    row += 1
mr_last_row = row - 1
mr.sheet_view.showGridLines = False

# data validation for Recognized? and Type
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
mr.add_data_validation(dv_yn)
dv_yn.add(f"G2:G{mr_last_row}")
dv_type = DataValidation(type="list", formula1='"Fee,Expense"', allow_blank=True)
mr.add_data_validation(dv_type)
dv_type.add(f"D2:D{mr_last_row}")

# =================================================================
# INVOICES - FINANCE
# =================================================================
inv = wb.create_sheet("Invoices - Finance")
set_widths(inv, [20, 30, 34, 16, 16, 22, 14, 40])
inv_headers = ["Invoice #", "Campaign", "SOW/Task ID(s)", "Period", "Amount",
               "Status", "Date Sent", "Notes for Finance"]
for i, h in enumerate(inv_headers, start=1):
    inv.cell(row=1, column=i, value=h)
style_header(inv, 1, len(inv_headers))
inv.freeze_panes = "A2"

# unique invoices with known numbers, plus the Haas F1 future invoices not yet earned (for full visibility)
invoice_list = [
    ("INV-38952AF", "CMP-05", "CMP5-SOW1", "Jan-26", "Sent to Finance", ""),
    ("INV-38953AF", "CMP-05", "CMP5-SOW1", "Feb-26 / Apr-26", "Sent to Finance", "Recognized across two months in source file — confirm with Finance this is one invoice."),
    ("INV-39386AF", "CMP-05", "CMP5-SOW1", "Mar-26", "Sent to Finance", ""),
    ("INV-39927AF", "CMP-05", "CMP5-SOW1", "Apr-26", "Sent to Finance", ""),
    ("INV-40021AF", "CMP-05", "CMP5-SOW1", "May-26", "Sent to Finance", ""),
    ("INV-40022AF", "CMP-05", "CMP5-SOW1", "Jun-26", "Not Yet Sent", "Not yet recognized — $15,000 open."),
    ("INV-40023AF", "CMP-05", "CMP5-SOW1", "Jul-26", "Not Yet Sent", "Not yet recognized — $17,250 open."),
    ("INV-40024AF", "CMP-05", "CMP5-SOW1", "Aug-26", "Not Yet Sent", "Not yet recognized — $21,500 open."),
    ("INV-40025AF", "CMP-05", "CMP5-SOW1", "Sep-26", "Not Yet Sent", "Not yet recognized — $14,000 open."),
    ("INV-40026AF", "CMP-05", "CMP5-SOW1", "Oct-26", "Not Yet Sent", "Not yet recognized — $14,000 open."),
    ("INV-40027AF", "CMP-05", "CMP5-SOW1", "Nov-26", "Not Yet Sent", "Not yet recognized — $21,500 open."),
    ("INV-40028AF", "CMP-05", "CMP5-SOW1", "Dec-26", "Not Yet Sent", "Not yet recognized — $21,500 open."),
    ("INV-37370AF", "CMP-07", "CMP7-SOW1", "Mar-26", "Sent to Finance", ""),
    ("INV-37371AF", "CMP-07", "CMP7-SOW1", "Apr-26", "Sent to Finance", ""),
    ("INV-37372AF", "CMP-07", "CMP7-SOW1", "May-26", "Sent to Finance", ""),
    ("INV-37373AF", "CMP-07", "CMP7-SOW1", "Jun-26", "Not Yet Sent", "Not yet recognized — $60,000 open."),
    ("INV-39804AO", "CMP-07", "CMP7-EXTRA1", "Mar-26", "Sent to Finance", "OOP Expenses."),
]

row = 2
for invoice, cid, sow, period, status, note in invoice_list:
    inv.cell(row=row, column=1, value=invoice).font = BLACK
    inv.cell(row=row, column=2, value=f'=IFERROR(INDEX(Campaigns!$B$2:$B$10,MATCH("{cid}",Campaigns!$A$2:$A$10,0)),"")').font = BLACK
    inv.cell(row=row, column=3, value=sow).font = BLACK
    inv.cell(row=row, column=4, value=period).font = BLUE
    inv.cell(row=row, column=5, value=f"=SUMIF('Monthly Recognition'!$F$2:$F$100,$A{row},'Monthly Recognition'!$E$2:$E$100)").font = BLACK
    inv.cell(row=row, column=5).number_format = CURRENCY
    inv.cell(row=row, column=6, value=status).font = BLUE
    inv.cell(row=row, column=7, value="").font = BLUE
    inv.cell(row=row, column=8, value=note).font = Font(name="Arial", size=9, italic=True, color="808080")
    if status == "Not Yet Sent":
        for c in range(1, 9):
            inv.cell(row=row, column=c).fill = FLAG_FILL
    border_row(inv, row, 8)
    row += 1
inv_last_row = row - 1

# revenue with no invoice # yet assigned — surfaced explicitly for Lisa to work through
row += 1
inv.cell(row=row, column=1, value="REVENUE WITH NO INVOICE # ASSIGNED YET").font = Font(name="Arial", bold=True, color="C00000")
inv.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
inv.cell(row=row, column=1, value=f'=SUMIF(\'Monthly Recognition\'!$F$2:$F$100,"Not yet assigned",\'Monthly Recognition\'!$E$2:$E$100)').font = Font(name="Arial", bold=True)
inv.cell(row=row, column=1).number_format = CURRENCY
inv.cell(row=row, column=2, value="Total recognized revenue across Brand Campaign Creative, Cancer Center Pavilion, U.S. News, Strategic Counsel, and part of Comms AOR that has never been tied to an invoice number. Assign invoice #s in Monthly Recognition column F to bring this to zero.").font = Font(name="Arial", italic=True)
inv.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)

dv_status = DataValidation(type="list", formula1='"Not Yet Sent,Sent to Finance,Confirmed by Finance"', allow_blank=True)
inv.add_data_validation(dv_status)
dv_status.add(f"F2:F{inv_last_row}")

inv.sheet_view.showGridLines = False

wb.save(os.path.join(SCRIPT_DIR, "Revenue Recognition Tracker.xlsx"))
print("saved")
